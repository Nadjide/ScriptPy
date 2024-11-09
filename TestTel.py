import tkinter as tk
from openpyxl.styles import PatternFill
from tkinterdnd2 import DND_FILES, TkinterDnD
import pandas as pd
import re
import openpyxl.utils
import os
from tkinter import messagebox
from tkinter import simpledialog


# Fonction pour vérifier si un numéro est valide
def est_valide(numero):
    numero = str(numero)
    numero = re.sub(r'\s+', '', numero)
    # Exemple de numéro valide: +33123456789 , +330123456789 , 123456789 , 33123456789 , 0612345678
    if (re.match(r'^\+33\d{9}$', numero) or re.match(r'^\+330\d{9}$', numero) or
            re.match(r'^0\d{9}$', numero) or re.match(r'^[1-9]\d{8}$', numero) or
            re.match(r'^33\d{9}$', numero)):
        return True, numero
    return False, numero

# Fonction pour demander les colonnes de téléphones à traiter

def demander_colonnes():
    colonnes = simpledialog.askstring(
        "Colonnes", "Entrez les lettres des colonnes contenant les numéros de téléphone (séparées par des virgules) :")
    colonnes = [col.strip().upper() for col in colonnes.split(',')]
    return [openpyxl.utils.column_index_from_string(col) - 1 for col in colonnes]

# Fonction pour traiter le fichier Excel
print("en cours")


def traiter_fichier(fichier, colonnes_telephone):
    df = pd.read_excel(fichier)
    
     # Vérifier si les colonnes existent dans le DataFrame
    colonnes_manquantes = [openpyxl.utils.get_column_letter(col + 1) for col in colonnes_telephone if df.columns[col] is None]
    if colonnes_manquantes:
        messagebox.showerror(
            "Colonnes manquantes", f"Les colonnes suivantes sont introuvables dans le fichier Excel : {', '.join(colonnes_manquantes)}")
        return
    
    num_telephones_invalides = 0
    num_doublons = 0
    num_telephones_vides = 0

    for col_idx in colonnes_telephone:
        col = df.columns[col_idx]
        # Ajouter une colonne temporaire avec les numéros corrigés
        df[f'{col}_corrige'] = df[col].apply(lambda x: est_valide(x)[1] if not pd.isna(x) else x)
        df[f'{col}_valide'] = df[f'{col}_corrige'].apply(lambda x: est_valide(x)[0])
        df[f'{col}_doublon'] = df[df[f'{col}_valide']].groupby(f'{col}_corrige')[f'{col}_corrige'].transform('count') > 1

    # Construire le chemin du fichier traité
    chemin_dossier, nom_fichier = os.path.split(fichier)
    nom_fichier_corrige = 'fichier_tel_corrige.xlsx'
    chemin_fichier_corrige = os.path.join(chemin_dossier, nom_fichier_corrige)

    # Création d'un nouveau fichier Excel avec openpyxl
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = 'Sheet1'

    # Copier les en-têtes de colonnes
    for col_num, column_title in enumerate(df.columns, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        new_ws[f'{col_letter}1'] = column_title

    # Copier les données et appliquer la mise en forme
    fill_invalides = PatternFill(
        start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    fill_doublons = PatternFill(
        start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    fill_vides = PatternFill(start_color='B7B7B7',
                             end_color='B7B7B7', fill_type='solid')

    for row_num, row_data in df.iterrows():
        for col_num, cell_value in enumerate(row_data, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            new_ws[f'{col_letter}{row_num + 2}'] = cell_value

            if (col_num - 1) in colonnes_telephone:
                if pd.isna(cell_value):
                    num_telephones_vides += 1
                    new_ws[f'{col_letter}{row_num + 2}'].fill = fill_vides
                else:
                    valide = df.loc[row_num, 'Valide']
                    doublon = df.loc[row_num, 'Doublon']
                    if not valide:
                        num_telephones_invalides += 1
                        new_ws[f'{col_letter}{row_num + 2}'].fill = fill_invalides
                    elif doublon:
                        num_doublons += 1
                        new_ws[f'{col_letter}{row_num + 2}'].fill = fill_doublons

    # Supprimer les colonnes temporaires
    del df['Telephone_corrige']
    del df['Valide']
    del df['Doublon']

    # Sauvegarder le nouveau fichier Excel
    new_wb.save(chemin_fichier_corrige)
    messagebox.showinfo(
        "Fichier créé", f"Le fichier traité a été enregistré sous : {chemin_fichier_corrige}")

    print(f"Nombre de téléphones invalides : {num_telephones_invalides}")
    print(f"Nombre de doublons : {num_doublons}")
    print(f"Nombre de téléphones vides : {num_telephones_vides}")

# Fonction pour gérer le glisser-déposer


def on_drop(event):
    fichier = event.data
    if isinstance(fichier, bytes):
        fichier = fichier.decode('utf-8')
    
    # Supprimer les caractères { et } s'ils sont présents dans le chemin du fichier
    fichier = fichier.strip('{}')
    
    print(f"Fichier reçu : {fichier}")

    if fichier.lower().endswith('.xlsx') or fichier.lower().endswith('.xlsm'):
        colonnes_telephone = demander_colonnes()
        traiter_fichier(fichier, colonnes_telephone)
    else:
        print("Extension de fichier incorrecte")
        print("Veuillez glisser un fichier .xlsx")



# Créer la fenêtre tkinter
fenetre = TkinterDnD.Tk()
fenetre.title("Traitement des numéros de téléphone")

# Configurer le glisser-déposer
fenetre.drop_target_register(DND_FILES)
fenetre.dnd_bind('<<Drop>>', on_drop)

# Définir les dimensions et la position de la fenêtre
fenetre.geometry("400x200")  # Largeur: 400, Hauteur: 200
screen_width = fenetre.winfo_screenwidth()
screen_height = fenetre.winfo_screenheight()
window_width = 400
window_height = 200
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)
fenetre.geometry(f"{window_width}x{window_height}+{x}+{y}")

# Ajouter un label pour afficher les instructions
label = tk.Label(
    fenetre, text="Glissez et déposez un fichier .xlsx ici", font=("Arial", 12))
label.pack(padx=10, pady=10)

# Afficher la fenêtre
fenetre.mainloop()
